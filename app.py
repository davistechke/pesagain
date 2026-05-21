import base64
import datetime
import functools
import os
import sqlite3
import uuid
import requests
import logging
import resend

from dotenv import load_dotenv
from flask import Flask, jsonify, redirect, render_template, request, session, url_for, make_response, flash
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from spin import spin_bp

load_dotenv()

logging.basicConfig(level=logging.DEBUG)

app = Flask(__name__)


app.wsgi_app = ProxyFix(
    app.wsgi_app,
    x_for=1,
    x_proto=1,
    x_host=1,
    x_prefix=1
)

app.secret_key = os.getenv("SECRET_KEY", "dev_secret_key_CHANGE_ME")

DB_PATH = os.getenv("DB_PATH", "app_database.db")

app.register_blueprint(spin_bp)

app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.getenv("RENDER") == "true"
app.config['SESSION_COOKIE_HTTPONLY'] = True

# =========================
# RESEND EMAIL CONFIG
# =========================
resend.api_key = os.getenv("RESEND_API_KEY")

serializer = URLSafeTimedSerializer(app.secret_key)

PAYHERO_BASE_URL   = os.getenv("PAYHERO_BASE_URL",   "https://backend.payhero.co.ke/api/v2")
PAYHERO_CHANNEL_ID = os.getenv("PAYHERO_CHANNEL_ID", "6532")
PAYHERO_PROVIDER   = os.getenv("PAYHERO_PROVIDER",   "m-pesa")
CALLBACK_URL       = ("https://gainpesaapp.onrender.com/callback" if os.getenv("RENDER")
                      else os.getenv("CALLBACK_URL", "https://cedrick-subdiscoid-drake.ngrok-free.de/callback"))
API_USERNAME       = os.getenv("API_USERNAME")
API_PASSWORD       = os.getenv("API_PASSWORD", "gMMRAHjO3snOZgQI7kS2xPpLlXLcylaKqaW5CJXd")

ACTIVATION_FEE         = 1.0
MIN_BINARY_DEPOSIT_KES = round(1.0 * 130.0, 2)


def get_auth_header():
    return f"Basic {base64.b64encode(f'{API_USERNAME}:{API_PASSWORD}'.encode()).decode()}"


def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            email TEXT PRIMARY KEY, username TEXT UNIQUE, password_hash TEXT, phone TEXT,
            balance REAL DEFAULT 0.0, spin_balance REAL DEFAULT 0.0,
            binary_balance REAL DEFAULT 0.0, binary_deposited REAL DEFAULT 0.0,
            binary_winnings REAL DEFAULT 0.0, total_earned REAL DEFAULT 0.0,
            total_withdrawn REAL DEFAULT 0.0, total_referred INTEGER DEFAULT 0,
            is_active BOOLEAN DEFAULT 0, referral_code TEXT UNIQUE,
            referred_by TEXT, joined_at TEXT, reset_token TEXT, token_expiry TEXT
        )
    """)
    for col, td in {"spin_balance":"REAL DEFAULT 0.0","binary_balance":"REAL DEFAULT 0.0",
                    "binary_deposited":"REAL DEFAULT 0.0","binary_winnings":"REAL DEFAULT 0.0",
                    "reset_token":"TEXT","token_expiry":"TEXT"}.items():
        try: conn.execute(f"ALTER TABLE users ADD COLUMN {col} {td}")
        except: pass

    conn.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            ext_ref TEXT PRIMARY KEY, email TEXT, type TEXT DEFAULT 'activation',
            status TEXT, amount REAL DEFAULT 0.0, FOREIGN KEY(email) REFERENCES users(email)
        )
    """)
    for col, td in [("type","TEXT DEFAULT 'activation'"),("amount","REAL DEFAULT 0.0")]:
        try: conn.execute(f"ALTER TABLE transactions ADD COLUMN {col} {td}")
        except: pass

    for ddl in [
        """CREATE TABLE IF NOT EXISTS withdrawals (
            id INTEGER PRIMARY KEY AUTOINCREMENT, email TEXT, amount REAL,
            mpesa_number TEXT, status TEXT, date TEXT, FOREIGN KEY(email) REFERENCES users(email))""",
        """CREATE TABLE IF NOT EXISTS binary_trades (
            id INTEGER PRIMARY KEY AUTOINCREMENT, email TEXT, asset TEXT, amount REAL,
            direction TEXT, status TEXT, payout REAL, timestamp TEXT,
            FOREIGN KEY(email) REFERENCES users(email))""",
        """CREATE TABLE IF NOT EXISTS admin_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT, admin_username TEXT,
            target_email TEXT, action_type TEXT, amount REAL, timestamp TEXT)""",
    ]:
        conn.execute(ddl)
    conn.commit()
    conn.close()


init_db()





def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if "user_email" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def send_reset_email(to_email: str, reset_link: str) -> bool:
    """Send password reset email via Resend."""
    try:
        params = {
            "from": "GainPesa <onboarding@resend.dev>",  # Change to your verified domain later e.g. noreply@gainpesa.com
            "to": [to_email],
            "subject": "GainPesa – Password Reset Request",
            "html": f"""
                <div style="font-family: Arial, sans-serif; max-width: 500px; margin: auto; padding: 20px;">
                    <h2 style="color: #2e7d32;">GainPesa Password Reset</h2>
                    <p>You requested a password reset. Click the button below to set a new password:</p>
                    <a href="{reset_link}" 
                       style="display:inline-block; background:#2e7d32; color:white; padding:12px 24px;
                              text-decoration:none; border-radius:6px; margin: 16px 0;">
                        Reset My Password
                    </a>
                    <p style="color:#666; font-size:13px;">This link expires in <strong>1 hour</strong>.</p>
                    <p style="color:#666; font-size:13px;">If you did not request this, ignore this email.</p>
                    <hr style="border:none; border-top:1px solid #eee; margin-top:30px;">
                    <p style="color:#aaa; font-size:11px;">GainPesa &copy; {datetime.datetime.now().year}</p>
                </div>
            """,
        }
        response = resend.Emails.send(params)
        app.logger.info(f"[RESEND] Email sent to {to_email} | Response: {response}")
        return True
    except Exception as e:
        app.logger.error(f"[RESEND ERROR] Failed to send to {to_email}: {e}")
        return False


def build_reset_url(token: str) -> str:
    return url_for("reset_password", token=token, _external=True)


# =========================
# DEBUG ROUTES
# =========================

@app.route("/debug-mail")
def debug_mail():
    return jsonify({
        "RESEND_API_KEY_SET": bool(os.getenv("RESEND_API_KEY")),
        "RENDER": os.getenv("RENDER"),
    })


@app.route("/test-mail")
def test_mail():
    try:
        params = {
            "from": "GainPesa <onboarding@resend.dev>",
            "to": ["delivered@resend.dev"],  # Resend's test address
            "subject": "GainPesa Test Email",
            "text": "Resend is working correctly on Render.",
        }
        response = resend.Emails.send(params)
        return jsonify({"success": True, "response": str(response)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/")
def index():
    return render_template("index.html")


@app.route('/manifest.json')
def manifest():
    return app.send_static_file('manifest.json')

@app.route('/sw.js')
def service_worker():
    return app.send_static_file('sw.js')


@app.route("/register", methods=["GET","POST"])
def register():
    error    = None
    ref_code = request.args.get("ref")
    if request.method == "POST":
        email       = request.form.get("email")
        username    = request.form.get("username")
        password    = request.form.get("password")
        phone       = request.form.get("phone")
        referred_by = request.form.get("ref")
        conn = get_db_connection()
        if conn.execute("SELECT 1 FROM users WHERE email=?", (email,)).fetchone():
            error = "Email already exists"
        elif conn.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
            error = "Username already taken"
        if error:
            conn.close()
            return render_template("register.html", error=error, ref_code=ref_code)
        conn.execute(
            "INSERT INTO users (email,username,password_hash,phone,referral_code,referred_by,joined_at) VALUES (?,?,?,?,?,?,?)",
            (email, username, generate_password_hash(password), phone,
             f"GP-{uuid.uuid4().hex.upper()[:5]}", referred_by or None,
             datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
        )
        conn.commit(); conn.close()
        session["user_email"] = email
        return redirect(url_for("pay_page"))
    return render_template("register.html", ref_code=ref_code)


@app.route("/login", methods=["GET","POST"])
def login():
    error = request.args.get("error")
    if request.method == "POST":
        credential = request.form.get("credential")
        password   = request.form.get("password")
        conn = get_db_connection()
        user = conn.execute("SELECT * FROM users WHERE email=? OR username=?", (credential, credential)).fetchone()
        conn.close()
        if not user or not check_password_hash(user["password_hash"], password):
            error = "Invalid credentials"
        else:
            session["user_email"] = user["email"]
            return redirect(url_for("dashboard") if user["is_active"] else url_for("pay_page"))
    return render_template("register.html", error=error)


@app.route("/forgot-password", methods=["GET","POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        conn  = get_db_connection()
        try:
            user = conn.execute("SELECT email FROM users WHERE LOWER(email)=?", (email,)).fetchone()
            if user:
                token      = serializer.dumps(user["email"], salt="gainpesa-password-reset")
                reset_link = build_reset_url(token)
                app.logger.info(f"[RESET LINK] Generated for {user['email']}: {reset_link}")
                email_sent = send_reset_email(user["email"], reset_link)
                if email_sent:
                    flash("Reset link sent — check your inbox (and spam folder).", "info")
                else:
                    flash("Could not send email. Please contact support.", "error")
            else:
                # Don't reveal whether email exists
                flash("If that email is registered, a reset link has been sent.", "info")
        except Exception as e:
            app.logger.error(f"[ForgotPassword CRITICAL] {e}", exc_info=True)
            flash("Something went wrong. Please try again.", "error")
        finally:
            conn.close()
        return redirect(url_for("forgot_password"))
    return render_template("forgot_password.html")


@app.route("/reset-password/<token>", methods=["GET","POST"])
def reset_password(token):
    try:
        email = serializer.loads(token, salt="gainpesa-password-reset", max_age=3600)
    except SignatureExpired:
        flash("Reset link expired (1-hour limit). Request a new one.", "error")
        return redirect(url_for("forgot_password"))
    except (BadSignature, Exception):
        flash("Invalid or already-used reset link. Request a new one.", "error")
        return redirect(url_for("forgot_password"))

    conn = get_db_connection()
    user = conn.execute("SELECT email FROM users WHERE email=?", (email,)).fetchone()
    if not user:
        conn.close()
        flash("Account not found.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":
        new_pw  = request.form.get("password", "")
        conf_pw = request.form.get("confirm_password", "")
        if len(new_pw) < 6:
            conn.close()
            return render_template("reset_password.html", token=token,
                                   error="Password must be at least 6 characters.")
        if new_pw != conf_pw:
            conn.close()
            return render_template("reset_password.html", token=token,
                                   error="Passwords do not match.")
        try:
            conn.execute("UPDATE users SET password_hash=? WHERE email=?",
                         (generate_password_hash(new_pw), email))
            conn.commit()
        except Exception as e:
            app.logger.error(f"[ResetPassword] DB error: {e}")
            conn.close()
            return render_template("reset_password.html", token=token,
                                   error="Could not save new password. Please try again.")
        conn.close()
        flash("✓ Password updated! You can now log in.", "success")
        return redirect(url_for("login"))

    conn.close()
    return render_template("reset_password.html", token=token)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/pay")
@login_required
def pay_page():
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE email=?", (session["user_email"],)).fetchone()
    conn.close()
    return render_template("pay.html", user=dict(user))


@app.route("/dashboard")
@login_required
def dashboard():
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE email=?", (session["user_email"],)).fetchone()
    conn.close()
    if not user["is_active"]: return redirect(url_for("pay_page"))
    return render_template("dashboard.html", user=dict(user))


@app.route("/gainbinary")
@login_required
def gainbinary():
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE email=?", (session["user_email"],)).fetchone()
    conn.close()
    if not user["is_active"]: return redirect(url_for("pay_page"))
    return render_template("gainbinary.html", user=dict(user))


@app.route("/api/initiate-payment", methods=["POST"])
@login_required
def initiate_payment():
    conn  = get_db_connection()
    phone = conn.execute("SELECT phone FROM users WHERE email=?", (session["user_email"],)).fetchone()["phone"]
    conn.close()
    if phone.startswith("0"): phone = "254"+phone[1:]
    elif phone.startswith("+"): phone = phone[1:]
    ext_ref = "GP-ACT-"+uuid.uuid4().hex[:6].upper()
    try:
        r = requests.post(f"{PAYHERO_BASE_URL}/payments", json={
            "amount":ACTIVATION_FEE,"phone_number":phone,"channel_id":PAYHERO_CHANNEL_ID,
            "provider":PAYHERO_PROVIDER,"external_reference":ext_ref,"callback_url":CALLBACK_URL
        }, headers={"Content-Type":"application/json","Authorization":get_auth_header()})
        if r.status_code in [200,201]:
            conn = get_db_connection()
            conn.execute("INSERT INTO transactions (ext_ref,email,type,status,amount) VALUES (?,?,?,?,?)",
                         (ext_ref,session["user_email"],"activation","pending",ACTIVATION_FEE))
            conn.commit(); conn.close()
            return jsonify({"success":True,"reference":ext_ref})
        return jsonify({"success":False,"error":r.text})
    except Exception as e:
        return jsonify({"success":False,"error":str(e)})


@app.route("/api/reconcile/<ext_ref>")
@login_required
def reconcile(ext_ref):
    conn = get_db_connection()
    tx   = conn.execute("SELECT status FROM transactions WHERE ext_ref=? AND email=?",
                        (ext_ref,session["user_email"])).fetchone()
    conn.close()
    if not tx: return jsonify({"status":"not_found"}),404
    return jsonify({"status":"confirmed" if tx["status"]=="confirmed"
                    else "canceled" if tx["status"]=="failed" else "pending"})


@app.route("/api/binary/deposit", methods=["POST"])
@login_required
def initiate_binary_deposit():
    amount = float(request.json.get("amount",0)); email = session["user_email"]
    if amount < MIN_BINARY_DEPOSIT_KES:
        return jsonify({"error":f"Minimum deposit is Ksh {MIN_BINARY_DEPOSIT_KES:.0f} (~1 USD)"}),400
    conn  = get_db_connection()
    phone = conn.execute("SELECT phone FROM users WHERE email=?", (email,)).fetchone()["phone"]
    conn.close()
    if phone.startswith("0"): phone = "254"+phone[1:]
    elif phone.startswith("+"): phone = phone[1:]
    ext_ref = "GP-BIN-"+uuid.uuid4().hex[:6].upper()
    try:
        r = requests.post(f"{PAYHERO_BASE_URL}/payments", json={
            "amount":amount,"phone_number":phone,"channel_id":PAYHERO_CHANNEL_ID,
            "provider":PAYHERO_PROVIDER,"external_reference":ext_ref,"callback_url":CALLBACK_URL
        }, headers={"Content-Type":"application/json","Authorization":get_auth_header()})
        if r.status_code in [200,201]:
            conn = get_db_connection()
            conn.execute("INSERT INTO transactions (ext_ref,email,type,status,amount) VALUES (?,?,?,?,?)",
                         (ext_ref,email,"binary_deposit","pending",amount))
            conn.commit(); conn.close()
            return jsonify({"success":True,"reference":ext_ref})
        return jsonify({"success":False,"error":r.text})
    except Exception as e:
        return jsonify({"success":False,"error":str(e)})


@app.route("/callback", methods=["POST"])
def callback():
    data=request.json; res=data.get("response") or data
    ext_ref=res.get("ExternalReference"); status=res.get("Status"); cb_amount=float(res.get("Amount",0))
    if not ext_ref: return jsonify({"status":"error"}),400
    conn=get_db_connection(); tx=conn.execute("SELECT * FROM transactions WHERE ext_ref=?",(ext_ref,)).fetchone()
    if not tx: conn.close(); return jsonify({"status":"not_found"}),404
    if str(status).lower() not in ["success","successful"]:
        conn.execute("UPDATE transactions SET status='failed' WHERE ext_ref=?",(ext_ref,))
        conn.commit(); conn.close(); return jsonify({"status":"ok"})
    tx_type=tx["type"] or "activation"; tx_amount=float(tx["amount"]) if tx["amount"] else cb_amount
    conn.execute("UPDATE transactions SET status='confirmed' WHERE ext_ref=?",(ext_ref,))
    if tx_type=="activation":
        conn.execute("UPDATE users SET is_active=1 WHERE email=?",(tx["email"],))
        bc=round(tx_amount,2)
        conn.execute("UPDATE users SET binary_balance=binary_balance+?,binary_deposited=binary_deposited+? WHERE email=?",(bc,bc,tx["email"]))
        ur=conn.execute("SELECT referred_by FROM users WHERE email=?",(tx["email"],)).fetchone()
        if ur and ur["referred_by"]:
            ref=conn.execute("SELECT email FROM users WHERE referral_code=?",(ur["referred_by"],)).fetchone()
            if ref:
                comm=round(tx_amount*0.50,2)
                conn.execute("UPDATE users SET balance=balance+?,total_earned=total_earned+?,total_referred=total_referred+1 WHERE email=?",(comm,comm,ref["email"]))
    elif tx_type=="binary_deposit":
        conn.execute("UPDATE users SET binary_balance=binary_balance+?,binary_deposited=binary_deposited+? WHERE email=?",(tx_amount,tx_amount,tx["email"]))
    conn.commit(); conn.close(); return jsonify({"status":"ok"})


@app.route("/api/binary/trade", methods=["POST"])
@login_required
def execute_binary_trade():
    data=request.json; email=session["user_email"]; amount=float(data.get("amount",0))
    conn=get_db_connection(); user=conn.execute("SELECT binary_balance FROM users WHERE email=?",(email,)).fetchone()
    if user["binary_balance"]<amount: conn.close(); return jsonify({"error":"Insufficient Trading Balance"}),400

    # 100% win rate: 80% profit on every trade
    payout = round(amount * 1.8, 2)
    profit = round(amount * 0.8, 2)
    conn.execute("UPDATE users SET binary_balance=binary_balance-?+? WHERE email=?",(amount,payout,email))
    conn.execute("UPDATE users SET binary_winnings=binary_winnings+?,total_earned=total_earned+? WHERE email=?",(payout,profit,email))
    conn.execute("INSERT INTO binary_trades (email,asset,amount,direction,status,payout,timestamp) VALUES (?,?,?,?,?,?,?)",
                 (email,data.get("asset","EUR/USD"),amount,data.get("direction"),"win",payout,datetime.datetime.now().strftime("%H:%M:%S")))
    conn.commit(); conn.close()
    return jsonify({"success":True,"status":"win","payout":payout,"profit":profit})


@app.route("/api/binary/claim-winnings", methods=["POST"])
@login_required
def claim_binary_winnings():
    email=session["user_email"]; amount=float(request.json.get("amount",0))
    conn=get_db_connection(); user=conn.execute("SELECT binary_winnings,binary_balance FROM users WHERE email=?",(email,)).fetchone()
    if amount<=0: conn.close(); return jsonify({"error":"Invalid amount"}),400
    if amount>round(user["binary_winnings"],2): conn.close(); return jsonify({"error":f"Available winnings: Ksh {user['binary_winnings']:.2f}."}),400
    if amount>user["binary_balance"]: conn.close(); return jsonify({"error":"Insufficient trading balance"}),400
    conn.execute("UPDATE users SET binary_balance=binary_balance-?,binary_winnings=binary_winnings-?,balance=balance+? WHERE email=?",(amount,amount,amount,email))
    conn.commit(); conn.close(); return jsonify({"success":True})


@app.route("/api/binary/transfer", methods=["POST"])
@login_required
def transfer_to_binary():
    amount=float(request.json.get("amount",0)); email=session["user_email"]
    conn=get_db_connection(); user=conn.execute("SELECT balance FROM users WHERE email=?",(email,)).fetchone()
    if user["balance"]<amount: conn.close(); return jsonify({"error":"Insufficient Wallet Balance"}),400
    conn.execute("UPDATE users SET balance=balance-?,binary_balance=binary_balance+?,binary_deposited=binary_deposited+? WHERE email=?",(amount,amount,amount,email))
    conn.commit(); conn.close(); return jsonify({"success":True})


@app.route("/api/user", methods=["GET"])
@login_required
def get_user_data():
    conn=get_db_connection()
    user=conn.execute("SELECT * FROM users WHERE email=?",(session["user_email"],)).fetchone()
    withdrawals=conn.execute("SELECT amount,mpesa_number as mpesa,status,date FROM withdrawals WHERE email=? ORDER BY id DESC",(session["user_email"],)).fetchall()
    conn.close()
    return jsonify({"balance":float(user["balance"] or 0),"binary_balance":float(user["binary_balance"] or 0),
        "binary_deposited":float(user["binary_deposited"] or 0),"binary_winnings":float(user["binary_winnings"] or 0),
        "withdrawable_balance":float(user["balance"] or 0),"total_earned":float(user["total_earned"] or 0),
        "total_withdrawn":float(user["total_withdrawn"] or 0),"total_referred":user["total_referred"],
        "referral_code":user["referral_code"],"min_binary_deposit":MIN_BINARY_DEPOSIT_KES,
        "withdrawals":[dict(w) for w in withdrawals]})


@app.route("/api/withdraw", methods=["POST"])
@login_required
def submit_withdraw():
    email=session["user_email"]; amount=float(request.json.get("amount",0)); mpesa=request.json.get("mpesa","")
    if amount<300: return jsonify({"error":"Minimum withdrawal is Ksh 300"}),400
    conn=get_db_connection()
    avail=round(float(conn.execute("SELECT balance FROM users WHERE email=?",(email,)).fetchone()["balance"] or 0),2)
    if amount>avail: conn.close(); return jsonify({"error":f"Only your earnings can be withdrawn. Available: Ksh {avail:.2f}"}),400
    conn.execute("UPDATE users SET balance=balance-?,total_withdrawn=total_withdrawn+? WHERE email=?",(amount,amount,email))
    conn.execute("INSERT INTO withdrawals (email,amount,mpesa_number,status,date) VALUES (?,?,?,?,?)",
                 (email,amount,mpesa,"pending",datetime.datetime.now().strftime("%b %d, %Y %H:%M")))
    conn.commit(); conn.close(); return jsonify({"success":True})


@app.route("/admin/login", methods=["GET","POST"])
def admin_login():
    if request.method=="POST":
        if request.form.get("username")=="MACK" and request.form.get("password")=="AJEGA":
            session["is_admin"]=True; return redirect(url_for("admin_dashboard"))
    return render_template("admin_login.html")


@app.route("/admin")
def admin_dashboard():
    if not session.get("is_admin"): return redirect(url_for("admin_login"))
    conn=get_db_connection()
    users=conn.execute("SELECT * FROM users ORDER BY joined_at DESC").fetchall()
    withdrawals=conn.execute("SELECT w.*,u.username FROM withdrawals w JOIN users u ON w.email=u.email ORDER BY w.id DESC").fetchall()
    recent_updates=conn.execute("SELECT l.*,u.username FROM admin_logs l JOIN users u ON l.target_email=u.email ORDER BY l.id DESC LIMIT 30").fetchall()
    conn.close()
    return render_template("admin.html",users=[dict(u) for u in users],
                           withdrawals=[dict(w) for w in withdrawals],recent_updates=[dict(r) for r in recent_updates])


@app.route("/admin/update-balance", methods=["POST"])
def admin_update_balance():
    if not session.get("is_admin"): return jsonify({"error":"Unauthorized"}),403
    email=request.json.get("email"); amt=float(request.json.get("balance",0))
    conn=get_db_connection()
    conn.execute("UPDATE users SET balance=balance+?,total_earned=total_earned+? WHERE email=?",(amt,amt,email))
    conn.execute("INSERT INTO admin_logs (admin_username,target_email,action_type,amount,timestamp) VALUES (?,?,?,?,?)",
                 ("MACK",email,"Wallet Addition",amt,datetime.datetime.now().strftime("%Y-%m-%d %H:%M")))
    conn.commit(); conn.close(); return jsonify({"success":True})


@app.route("/admin/update-trading", methods=["POST"])
def admin_update_trading():
    if not session.get("is_admin"): return jsonify({"error":"Unauthorized"}),403
    email=request.json.get("email"); amt=float(request.json.get("amount",0))
    conn=get_db_connection()
    conn.execute("UPDATE users SET binary_balance=binary_balance+? WHERE email=?",(amt,email))
    conn.execute("INSERT INTO admin_logs (admin_username,target_email,action_type,amount,timestamp) VALUES (?,?,?,?,?)",
                 ("MACK",email,"Binary Addition",amt,datetime.datetime.now().strftime("%Y-%m-%d %H:%M")))
    conn.commit(); conn.close(); return jsonify({"success":True})


@app.route("/admin/mark-paid", methods=["POST"])
def admin_mark_paid():
    if not session.get("is_admin"): return jsonify({"error":"Unauthorized"}),403
    conn=get_db_connection()
    conn.execute("UPDATE withdrawals SET status='paid' WHERE id=?",(request.json.get("id"),))
    conn.commit(); conn.close(); return jsonify({"success":True})


@app.route("/admin/download-pdf/<status>")
def download_users_pdf(status):
    if not session.get("is_admin"): return redirect(url_for("admin_login"))
    conn = get_db_connection()
    if status == "activated":
        users = conn.execute("SELECT * FROM users WHERE is_active=1").fetchall()
    elif status == "pending":
        users = conn.execute("SELECT * FROM users WHERE is_active=0").fetchall()
    else:
        users = conn.execute("SELECT * FROM users").fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{status.title()} Users"

    # ── Title row ──────────────────────────────────────────────────
    green  = "1B5E20"
    lgreen = "C8E6C9"
    white  = "FFFFFF"
    grey   = "F5F5F5"

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"GAINPESA — {status.upper()} USERS REPORT"
    title_cell.font      = Font(name="Arial", bold=True, size=14, color=white)
    title_cell.fill      = PatternFill("solid", start_color=green)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Generated date row
    ws.merge_cells("A2:G2")
    date_cell = ws["A2"]
    date_cell.value = f"Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}"
    date_cell.font      = Font(name="Arial", italic=True, size=9, color="555555")
    date_cell.alignment = Alignment(horizontal="right")
    ws.row_dimensions[2].height = 16

    # ── Header row ─────────────────────────────────────────────────
    headers = ["#", "Email", "Username", "Phone", "Balance (Ksh)",
               "Status", "Joined At"]
    thin = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color=white)
        cell.fill      = PatternFill("solid", start_color=green)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[3].height = 20

    # ── Data rows ──────────────────────────────────────────────────
    for row_idx, u in enumerate(users, start=1):
        excel_row = row_idx + 3
        fill_color = lgreen if row_idx % 2 == 0 else grey
        row_data = [
            row_idx,
            u["email"],
            u["username"],
            u["phone"] or "",
            round(float(u["balance"] or 0), 2),
            "Active" if u["is_active"] else "Pending",
            u["joined_at"] or "",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = PatternFill("solid", start_color=fill_color)
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 5, 6] else "left",
                                       vertical="center")
            cell.border    = border
            if col_idx == 5:
                cell.number_format = '#,##0.00'
            if col_idx == 6:
                cell.font = Font(name="Arial", size=9,
                                 color="1B5E20" if u["is_active"] else "B71C1C",
                                 bold=True)

    # ── Totals row ─────────────────────────────────────────────────
    total_row = len(users) + 4
    ws.cell(row=total_row, column=4, value="TOTAL BALANCE").font = Font(bold=True, name="Arial", size=9)
    total_cell = ws.cell(row=total_row, column=5,
                         value=f"=SUM(E4:E{total_row - 1})")
    total_cell.font         = Font(bold=True, name="Arial", size=9, color=green)
    total_cell.number_format = '#,##0.00'
    total_cell.border        = border

    # ── Column widths ──────────────────────────────────────────────
    col_widths = [5, 35, 18, 18, 16, 12, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze panes below header ──────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Stream to response ─────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resp = make_response(output.read())
    resp.headers.set("Content-Disposition", "attachment",
                     filename=f"{status}_users.xlsx")
    resp.headers.set("Content-Type",
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return resp


@app.route("/admin/activate-user", methods=["POST"])
def admin_activate_user():
    if not session.get("is_admin"): return jsonify({"error":"Unauthorized"}),403
    email=request.json.get("email"); conn=get_db_connection()
    conn.execute("UPDATE users SET is_active=1 WHERE email=?",(email,))
    bc=round(ACTIVATION_FEE,2)
    conn.execute("UPDATE users SET binary_balance=binary_balance+?,binary_deposited=binary_deposited+? WHERE email=?",(bc,bc,email))
    ur=conn.execute("SELECT referred_by FROM users WHERE email=?",(email,)).fetchone()
    if ur and ur["referred_by"]:
        ref=conn.execute("SELECT email FROM users WHERE referral_code=?",(ur["referred_by"],)).fetchone()
        if ref:
            comm=round(ACTIVATION_FEE*0.50,2)
            conn.execute("UPDATE users SET balance=balance+?,total_earned=total_earned+?,total_referred=total_referred+1 WHERE email=?",(comm,comm,ref["email"]))
    conn.execute("INSERT INTO admin_logs (admin_username,target_email,action_type,amount,timestamp) VALUES (?,?,?,?,?)",
                 ("MACK",email,"Manual Activation",ACTIVATION_FEE,datetime.datetime.now().strftime("%Y-%m-%d %H:%M")))
    conn.commit(); conn.close(); return jsonify({"success":True})


if __name__ == "__main__":
    app.run(debug=True)
